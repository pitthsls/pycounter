"""COUNTER journal and book reports and associated functions."""

import collections
import datetime
import logging
import re
import warnings

import pendulum

from pycounter import csvhelper
from pycounter.constants import CODES, HEADER_FIELDS, METRICS
from pycounter.constants import REPORT_DESCRIPTIONS, TOTAL_TEXT
from pycounter.exceptions import (
    PycounterException,
    PycounterWarning,
    UnknownReportTypeError,
)
from pycounter.helpers import (
    convert_covered,
    convert_date_run,
    format_stat,
    guess_type_from_content,
    is_first_last,
    next_month,
)


class CounterReport:
    """
    a COUNTER usage statistics report.

    Iterate over the report object to get its rows (each of which is a
    :class:`CounterBook <CounterBook>` or :class:`CounterJournal
    <CounterJournal>` instance.

    :param metric: metric being tracked by this report. For database
        reports (which have multiple metrics per report), this should be
        set to `None`.

    :param report_type: type of report (e.g., "JR1", "BR2")

    :param report_version: COUNTER version

    :param customer: name of customer on report

    :param institutional_identifier: unique ID assigned by vendor for
        customer

    :param period: tuple of datetime.date objects corresponding to the
        beginning and end of the covered range

    :param date_run: date the COUNTER report was generated

    :param section_type: predominant section type used for this report.
        (applies to report BR2; should probably be None for any other report
        type)

    """

    # pylint: disable=too-many-instance-attributes

    def __init__(
        self,
        report_type=None,
        report_version=4,
        metric=None,
        customer=None,
        institutional_identifier=None,
        period=(None, None),
        date_run=None,
        section_type=None,
    ):
        self.pubs = []
        self.report_type = report_type
        self.report_version = report_version
        self.metric = metric
        self.customer = customer
        self.institutional_identifier = institutional_identifier
        if not is_first_last(period):
            warnings.warn(
                "report period should be from"
                "first day of a month to last day of a month.",
                PycounterWarning,
            )
        self.period = period
        if date_run is None:
            self.date_run = datetime.date.today()
        else:
            self.date_run = date_run
        self.section_type = section_type

    def __repr__(self):
        return "<CounterReport {} version {} for date range {} to {}>".format(
            self.report_type, self.report_version, self.period[0], self.period[1]
        )

    def __iter__(self):
        return iter(self.pubs)

    def write_to_file(self, path, format_):
        """
        Output report to a file.

        :param path: location to write file
        :param format_: file format. Currently supports 'tsv'
        :return:
        """
        if format_ == "tsv":
            self.write_tsv(path)
        else:
            raise PycounterException("unknown file type %s" % format_)

    def write_tsv(self, path):
        """
        Output report to a COUNTER 4 TSV file.

        :param path: location to write file
        """
        lines = self.as_generic()
        with csvhelper.UnicodeWriter(path, delimiter="\t") as writer:
            writer.writerows(lines)

    def as_generic(self):
        """
        Output report as list of lists.

        Nested list will contain cells that would appear
        in COUNTER report (suitable for writing as CSV, TSV, etc.)
        """
        output_lines = []
        rep_type = ""
        for name, code in CODES.items():
            if code == self.report_type[0:2]:
                rep_type = name

        report_name = "{} Report {} (R{})".format(
            rep_type, self.report_type[-1], self.report_version
        )
        output_lines.append([report_name, REPORT_DESCRIPTIONS[self.report_type]])
        if self.report_type == "BR2":
            output_lines.append([self.customer, "Section Type:"])
            output_lines.append([self.institutional_identifier, self.section_type])
        else:
            output_lines.append([self.customer])
            output_lines.append([self.institutional_identifier])
        output_lines.append(["Period covered by Report:"])
        period = "{} to {}".format(
            self.period[0].strftime("%Y-%m-%d"), self.period[1].strftime("%Y-%m-%d")
        )
        output_lines.append([period])
        output_lines.append(["Date run:"])
        output_lines.append([self.date_run.strftime("%Y-%m-%d")])
        output_lines.append(self._table_header())
        if self.report_type in ("JR1", "BR1", "BR2", "DB2", "JR2", "BR3"):
            output_lines.extend(self._totals_lines())
        elif self.report_type.startswith("DB"):
            self._ensure_required_metrics()
            try:
                self.pubs.sort(key=lambda x: METRICS[self.report_type].index(x.metric))
            except ValueError:  # pragma: nocover
                pass

        for pub in sorted(self.pubs, key=lambda x: x.title or ''):
            output_lines.append(pub.as_generic())

        return output_lines

    def _totals_lines(self):
        """Generate Totals for COUNTER report, as list of lists of cells."""
        total_lines = []
        metrics = {resource.metric for resource in self.pubs}

        for metric in sorted(metrics):
            total_lines.append(self._totals_line(metric))

        return total_lines

    def _totals_line(self, metric):
        """Generate Totals for a given metric."""
        total_cells = [TOTAL_TEXT[self.report_type]]
        publishers = {resource.publisher for resource in self.pubs}
        if len(publishers) == 1:
            total_cells.append(publishers.pop())
        else:
            total_cells.append("")
        platforms = {resource.platform for resource in self.pubs}
        if len(platforms) == 1:
            total_cells.append(platforms.pop())
        else:
            total_cells.append("")
        if self.report_type in ("JR1", "BR1", "BR2", "JR2", "BR3"):
            total_cells.extend([""] * 4)
        if self.report_type in ("DB2", "JR2", "BR3"):
            total_cells.append(metric)
        total_usage = 0
        pdf_usage = 0
        html_usage = 0

        start_month_first_day = datetime.date(
            self.period[0].year, self.period[0].month, 1
        )
        months = list(
            pendulum.Period(start_month_first_day, self.period[1]).range("months")
        )
        month_data = [0] * len(months)
        for pub in self.pubs:
            if pub.metric != metric:
                continue
            if self.report_type == "JR1":
                pdf_usage += pub.pdf_total  # pytype: disable=attribute-error
                html_usage += pub.html_total  # pytype: disable=attribute-error
            for data in pub:
                total_usage += data[2]
                month_data[months.index(data[0])] += data[2]
        total_cells.append(str(total_usage))
        if self.report_type == "JR1":
            total_cells.append(str(html_usage))
            total_cells.append(str(pdf_usage))
        total_cells.extend(str(d) for d in month_data)
        return total_cells

    def _table_header(self):
        """Generate header for COUNTER table for report, as list of cells."""
        header_cells = list(HEADER_FIELDS[self.report_type])
        start_month_first_day = datetime.date(
            self.period[0].year, self.period[0].month, 1
        )
        for d_obj in pendulum.Period(start_month_first_day, self.period[1]).range(
            "months"
        ):
            header_cells.append(d_obj.strftime("%b-%Y"))
        return header_cells

    def _ensure_required_metrics(self):
        """
        Build up a dict of sets of known metrics for each database.

        If any metric is missing add a 0 use
        :class:`CounterDatabase<CounterDatabase>`.
        Assumes platform and publisher are consistent across records.
        """
        try:
            required_metrics = METRICS[self.report_type]
        except LookupError:  # pragma: nocover
            raise UnknownReportTypeError(self.report_type)

        dbs = collections.defaultdict(set)
        for database in self.pubs:
            dbs[database.title].add(database.metric)

        for database, metrics in dbs.items():
            for metric in (m for m in required_metrics if m not in metrics):
                self.pubs.append(
                    CounterDatabase(
                        title=database,
                        platform=self.pubs[0].platform,
                        publisher=self.pubs[0].publisher,
                        period=self.period,
                        metric=metric,
                        month_data=[(self.period[0], 0)],
                    )
                )


MonthsUsage = collections.namedtuple("MonthsUsage", "month metric usage")


class CounterEresource:
    """
    Base class for COUNTER statistics lines.

    Iterating returns (first_day_of_month, metric, usage) tuples.

    :param period: two-tuple of datetime.date objects corresponding
        to the beginning and end dates of the covered range

    :param metric: metric tracked by this report. Should be a value
        from pycounter.report.METRICS dict.

    :param month_data: a list containing usage data for this
        resource, as (datetime.date, usage) tuples

    :param title: title of the resource

    :param publisher: name of the resource's publisher

    :param platform: name of the platform providing the resource

    """

    def __init__(
        self,
        period=None,
        metric=None,
        month_data=None,
        title="",
        platform="",
        publisher="",
    ):
        self.period = period

        self.metric = metric
        self._full_data = []
        if month_data is not None:
            for item in month_data:
                self._full_data.append(item)

        self.title = title
        self.platform = platform
        self.publisher = publisher

    def __iter__(self):
        if self._full_data:
            for item in self._full_data:
                yield MonthsUsage(item[0], self.metric, item[1])

    def _fill_months(self):
        """Ensure each month in period represented and zero fill if not."""
        start_month_first_day = datetime.date(
            self.period[0].year, self.period[0].month, 1
        )
        start, end = start_month_first_day, self.period[1]
        try:
            for d_obj in pendulum.Period(start, end).range("months"):
                if d_obj not in (x[0] for x in self._full_data):
                    self._full_data.append((d_obj, 0))
        except IndexError:  # pragma: nocover
            pass
        else:
            self._full_data.sort()


class CounterJournal(CounterEresource):
    """
    Statistics for a single electronic journal.

    :param period: two-tuple of datetime.date objects corresponding
        to the beginning and end dates of the covered range

    :param metric: the metric tracked by this statistics line.
        (Should probably always be "FT Article Requests" for
        CounterJournal objects, as long as only JR1 is supported.)

    :param issn: eJournal's print ISSN

    :param eissn: eJournal's eISSN

    :param month_data: a list containing usage data for this
        journal, as (datetime.date, usage) tuples

    :param title: title of the resource

    :param publisher: name of the resource's publisher

    :param platform: name of the platform providing the resource

    :param html_total: total HTML usage for this title for reporting period

    :param pdf_total: total PDF usage for this title for reporting period

    """

    def __init__(
        self,
        period=None,
        metric=METRICS["JR1"],
        issn=None,
        eissn=None,
        month_data=None,
        title="",
        platform="",
        publisher="",
        html_total=0,
        pdf_total=0,
        doi="",
        proprietary_id="",
    ):
        super().__init__(period, metric, month_data, title, platform, publisher)
        self.html_total = html_total
        self.pdf_total = pdf_total
        self.doi = doi
        self.proprietary_id = proprietary_id

        self.isbn = None

        if issn is not None:
            self.issn = issn
        else:
            self.issn = ""

        if eissn is not None:
            self.eissn = eissn
        else:
            self.eissn = ""

    def __repr__(self):  # pragma: nocover
        return """<CounterJournal %s, publisher %s,
        platform %s>""" % (
            self.title,
            self.publisher,
            self.platform,
        )

    def as_generic(self):
        """Get data for this line as list of COUNTER report cells."""
        self._fill_months()  # Ensure fill all months with zero at least
        data_line = [
            self.title,
            self.publisher,
            self.platform,
            self.doi,
            self.proprietary_id,
            self.issn,
            self.eissn,
        ]
        total_usage = 0
        month_data = []
        for data in self:
            total_usage += data[2]
            month_data.append(str(data[2]))
        if self.metric.startswith("Access"):
            data_line.append(self.metric)
        data_line.append(str(total_usage))
        if not self.metric.startswith("Access"):
            data_line.append(str(self.html_total))
            data_line.append(str(self.pdf_total))
        data_line.extend(month_data)
        return data_line


class CounterBook(CounterEresource):
    """
    statistics for a single electronic book.

    :ivar isbn: eBook's ISBN

    :ivar issn: eBook's ISSN (if any)

    :param month_data: a list containing usage data for this
        book, as (datetime.date, usage) tuples

    :param title: title of the resource

    :param publisher: name of the resource's publisher

    :param platform: name of the platform providing the resource

    """

    def __init__(
        self,
        period=None,
        metric=None,
        month_data=None,
        title="",
        platform="",
        publisher="",
        isbn=None,
        issn=None,
        doi="",
        proprietary_id="",
        print_isbn=None,
        online_isbn=None,
    ):
        super().__init__(period, metric, month_data, title, platform, publisher)
        self.eissn = None
        self.doi = doi
        self.proprietary_id = proprietary_id

        self._isbn = isbn
        self.print_isbn = print_isbn
        self.online_isbn = online_isbn

        if issn is not None:
            self.issn = issn
        else:
            self.issn = ""

    def __repr__(self):
        return """<CounterBook %s (ISBN: %s), publisher %s,
        platform %s>""" % (
            self.title,
            self.isbn,
            self.publisher,
            self.platform,
        )

    @property
    def isbn(self):
        """Return a suitable ISSN for the ebook.

        The tabular COUNTER reports only report an "ISBN", while the SUSHI
        (XML) reports include both a Print_ISBN and Online_ISBN.

         This property will return a generic ISBN given in the constructor,
         if any. If the CounterBook was created with no "isbn" but with
         online_ISBN and/or print_ISBN, the online one, if any, will be
         returned, otherwise the print.
        """
        return self._isbn or self.online_isbn or self.print_isbn or ""

    def as_generic(self):
        """Get data for this line as list of COUNTER report cells."""
        self._fill_months()  # Ensure fill all months with zero at least
        data_line = [
            self.title,
            self.publisher,
            self.platform,
            self.doi,
            self.proprietary_id,
            self.isbn,
            self.issn,
        ]
        total_usage = 0
        month_data = []
        for data in self:
            total_usage += data[2]
            month_data.append(str(data[2]))
        if self.metric and self.metric.startswith("Access"):
            data_line.append(self.metric)
        data_line.append(str(total_usage))
        data_line.extend(month_data)
        return data_line


class CounterDatabase(CounterEresource):
    """a COUNTER database report line."""

    def __init__(
        self,
        period=None,
        metric=None,
        month_data=None,
        title="",
        platform="",
        publisher="",
    ):
        super().__init__(period, metric, month_data, title, platform, publisher)
        self.isbn = None

    def as_generic(self):
        """Return data for this line as list of COUNTER report cells."""
        self._fill_months()

        data_line = [self.title, self.publisher, self.platform, self.metric]
        total_usage = 0
        month_data = []

        for data in self:
            total_usage += data[2]
            month_data.append(str(data[2]))

        data_line.append(str(total_usage))
        data_line.extend(month_data)

        return data_line


class CounterPlatform(CounterEresource):
    """a COUNTER platform report line."""

    def __init__(
        self, period=None, metric=None, month_data=None, platform="", publisher=""
    ):
        super().__init__(
            period=period,
            metric=metric,
            month_data=month_data,
            title="",  # no title for platform report
            platform=platform,
            publisher=publisher,
        )
        self.isbn = None

    def as_generic(self):
        """Return data for this line as list of COUNTER report cells."""
        self._fill_months()

        data_line = [self.platform, self.publisher, self.metric]
        total_usage = 0
        month_data = []

        for data in self:
            total_usage += data[2]
            month_data.append(str(data[2]))

        data_line.append(str(total_usage))
        data_line.extend(month_data)

        return data_line


def parse(filename, filetype=None, encoding="utf-8", fallback_encoding="latin-1"):
    """Parse a COUNTER file, first attempting to determine type.

    Returns a :class:`CounterReport <CounterReport>` object.

    :param filename: path to COUNTER report to load and parse.
    :param filetype: type of file provided, one of "csv", "tsv", "xlsx".
        If set to None (the default), an attempt will be made to
        detect the correct type, first from the file extension, then from
        the file's contents.
    :param encoding: encoding to use to decode the file. Defaults to 'utf-8',
        ignored for XLSX files (which specify their encoding in their XML)
    :param fallback_encoding: alternative encoding to use to try to decode
        the file if the primary encoding fails. This defaults to 'latin-1',
        which will accept any bytes (possibly producing junk results...)
        Ignored for XLSX files.

    """
    if filetype is None:
        if filename.endswith(".tsv"):
            filetype = "tsv"
        elif filename.endswith(".xlsx"):
            filetype = "xlsx"
        elif filename.endswith(".csv"):
            filetype = "csv"
        else:
            with open(filename, "rb") as file_obj:
                filetype = guess_type_from_content(file_obj)

    if filetype == "tsv":
        return parse_separated(filename, "\t", encoding, fallback_encoding)
    if filetype == "xlsx":
        return parse_xlsx(filename)
    if filetype == "csv":
        return parse_separated(filename, ",", encoding, fallback_encoding)
    raise PycounterException("Unknown file type %s" % filetype)


def parse_xlsx(filename):
    """Parse a COUNTER file in Excel format.

    Invoked automatically by ``parse``.

    :param filename: path to XLSX-format COUNTER report file.

    """
    from openpyxl import load_workbook  # pylint: disable=import-outside-toplevel

    with open(filename, "rb") as xlsx_file:
        workbook = load_workbook(xlsx_file)
        worksheet = workbook[workbook.sheetnames[0]]
        row_it = worksheet.iter_rows()
        split_row_list = (
            [cell.value if cell.value is not None else "" for cell in row]
            for row in row_it
        )

    return parse_generic(split_row_list)


def parse_separated(filename, delimiter, encoding="utf-8", fallback_encoding="latin-1"):
    r"""Open COUNTER CSV/TSV report and parse into a CounterReport.

    Invoked automatically by :py:func:`parse`.

    :param filename: path to delimited COUNTER report file.

    :param delimiter: character (such as ',' or '\\t') used as the
        delimiter for this file

    :param encoding: file's encoding. Default: utf-8

    :param fallback_encoding: alternative encoding to try to decode if
        default fails. Throws a warning if used.

    :return: CounterReport object
    """
    with csvhelper.UnicodeReader(
        filename,
        delimiter=delimiter,
        fallback_encoding=fallback_encoding,
        encoding=encoding,
    ) as report_reader:
        return parse_generic(report_reader)


def parse_generic(report_reader):
    """Parse COUNTER report rows into a CounterReport.

    :param report_reader: a iterable object that yields lists COUNTER
        data formatted as tabular lists
    :return: CounterReport object

    """
    # pylint: disable=too-many-branches
    report = CounterReport()

    first_line = next(report_reader)
    if first_line[0] == "Report_Name":  # COUNTER 5 report
        second_line = next(report_reader)
        third_line = next(report_reader)
        report.report_type, report.report_version = _get_c5_type_and_version(
            second_line, third_line
        )
    else:
        report.report_type, report.report_version = _get_type_and_version(first_line[0])

    if report.report_version != 5:
        # noinspection PyTypeChecker
        report.metric = METRICS.get(report.report_type)

    report.customer = next(report_reader)[1 if report.report_version == 5 else 0]

    if report.report_version >= 4:
        inst_id_line = next(report_reader)
        if inst_id_line:
            report.institutional_identifier = inst_id_line[
                1 if report.report_version == 5 else 0
            ]
            if report.report_type == "BR2":
                report.section_type = inst_id_line[1]

        next(report_reader)
        if report.report_version == 5:
            for _ in range(3):
                next(report_reader)

        covered_line = next(report_reader)
        report.period = convert_covered(
            covered_line[1 if report.report_version == 5 else 0]
        )

    if report.report_version < 5:
        next(report_reader)

    date_run_line = next(report_reader)
    report.date_run = convert_date_run(
        date_run_line[1 if report.report_version == 5 else 0]
    )

    if report.report_version == 5:
        for _ in range(2):
            # Skip Created_By and blank line
            next(report_reader)

    header = next(report_reader)

    countable_header = header[0:8]
    for col in header[8:]:
        if col:
            countable_header.append(col)
    last_col = len(countable_header)

    if report.report_type not in ("DB1", "PR1") and report.report_version != 5:
        # these reports do not have line with totals
        next(report_reader)

    if report.report_type in ("DB2", "BR3", "JR3"):
        # this report has two lines of totals
        next(report_reader)

    for line in report_reader:
        if not line:
            continue
        report.pubs.append(_parse_line(line, report, last_col))

    return report


def _parse_line(line, report, last_col):
    """Parse a single line from a report.

    :param line: sequence of cells in a report line
    :param report: a CounterReport the line came from
    :param last_col: last column number containing data
    :return: an appropriate CounterResource subclass instance
    """
    # pylint: disable=too-many-locals
    issn = None
    eissn = None
    isbn = None
    html_total = 0
    pdf_total = 0
    doi = ""
    prop_id = ""

    metric = report.metric
    if (
        report.report_type.startswith("JR1")
        or report.report_type == "TR_J2"
    ):

        old_line = line
        line = line[0:3] + line[5:7] + line[10:last_col]
        doi = old_line[3]
        prop_id = old_line[4]
        html_total = format_stat(old_line[8])
        pdf_total = format_stat(old_line[9])
        issn = line[3].strip()
        eissn = line[4].strip()

    elif report.report_type in "TR_J1":
        old_line = line
        line = line[0:2] + line[3:4] + line[6:8] + line[11:last_col]
        doi = old_line[4]
        metric = old_line[9]
        prop_id = old_line[5]
        issn = line[3].strip()
        eissn = line[4].strip()

    elif report.report_type in ("BR1", "BR2"):
        line = line[0:3] + line[5:7] + line[8:last_col]
        isbn = line[3].strip()
        issn = line[4].strip()

    elif report.report_type in ("BR3", "JR2"):
        metric = line[7]
        doi = line[3]
        prop_id = line[4]
        line = line[0:3] + line[5:7] + line[9:last_col]
        eissn = line[4].strip()
        if report.report_type == "BR3":
            isbn = line[3].strip()
        else:
            issn = line[3].strip()
    # For DB1 and DB2, nothing additional to do here

    logging.debug(line)
    common_args = {
        "title": line[0],
        "publisher": line[1],
        "platform": line[2],
        "period": report.period,
    }
    month_data = []
    curr_month = datetime.date(report.period[0].year, report.period[0].month, 1)
    months_start_idx = 5 if report.report_type != "PR1" else 4
    for data in line[months_start_idx:]:
        month_data.append((curr_month, format_stat(data)))
        curr_month = next_month(curr_month)
    if (
        report.report_type.startswith("JR")
        or report.report_type == "TR_J1"
        or report.report_type == "TR_J2"
    ):
        return CounterJournal(
            metric=metric,
            month_data=month_data,
            doi=doi,
            issn=issn,
            eissn=eissn,
            proprietary_id=prop_id,
            html_total=html_total,
            pdf_total=pdf_total,
            **common_args
        )
    if report.report_type.startswith("BR"):
        return CounterBook(
            metric=metric,
            month_data=month_data,
            doi=doi,
            issn=issn,
            isbn=isbn,
            proprietary_id=prop_id,
            **common_args
        )
    if report.report_type.startswith("DB"):
        return CounterDatabase(metric=line[3], month_data=month_data, **common_args)
    if report.report_type == "PR1":
        # there is no title in the PR1 report
        return CounterPlatform(
            metric=line[2],
            month_data=month_data,
            platform=line[0],
            publisher=line[1],
            period=report.period,
        )
    raise PycounterException("Should be unreachable")  # pragma: no cover


def _get_type_and_version(specifier):
    """Given a COUNTER report specifier, find the type and version.

    :param specifier: COUNTER report specifier
    :return: type, version tuple
    """
    report_types_clause = "|".join(CODES)
    rt_match = re.match(
        r".*(%s) Report (\d(?: GOA)?) ?\(R(\d)\)" % report_types_clause, specifier
    )
    if rt_match:
        report_type = CODES[rt_match.group(1)] + rt_match.group(2)
        report_version = int(rt_match.group(3))
    else:
        raise UnknownReportTypeError("No match in line: %s" % specifier)
    # pragma: nocover
    if not any(report_type.startswith(x) for x in ("JR", "BR", "DB", "PR1")):
        raise UnknownReportTypeError(report_type)

    if report_version < 4:  # pragma: nocover
        raise UnknownReportTypeError("Only COUNTER 4&5 are supported.")

    return report_type, report_version


def _get_c5_type_and_version(second_line, third_line):
    """Find COUNTER 5 specific type and version."""
    return second_line[1], int(third_line[1])


def _year_from_header(header, report):
    """Get the year of the first month in the report from the header.

    (Only used for COUNTER 4.)
    """
    first_date_col = 10 if report.report_version == 4 else 5
    if report.report_type in ("BR1", "BR2") and report.report_version == 4:
        first_date_col = 8
    elif report.report_type in ("DB1", "DB2") and report.report_version == 4:
        first_date_col = 5
    elif report.report_type == "PR1" and report.report_version == 4:
        first_date_col = 4
    elif report.report_type == "JR2":
        first_date_col = 9
    year = int(header[first_date_col].split("-")[1])
    if year < 100:
        year += 2000

    return year
